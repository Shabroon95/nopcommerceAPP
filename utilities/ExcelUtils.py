import openpyxl
from openpyxl.styles import PatternFill

def getRowCount(file, sheetName):
    workbook=openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return(sheet.max_row)

def getColumnCount(file, sheetname):
    workbook=openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return (sheet.max_column)

def readData(file,sheetName,rownum,columnno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.cell(rownum, columnno).value

def writeData(file,sheetName, rownum,columnno,Data):
    workbook= openpyxl.workbook(file)
    sheet= workbook[sheetName]
    sheet.cell(rownum,columnno).value= Data
    workbook.save(file)

def fillGreenColor(file,sheeName, rowno, columnno):
    workbook=openpyxl.load_workbook(file)
    sheet= workbook[sheeName]
    greenfill = PatternFill(start_color='60b212', end_color='60b212',fill_type='solid')
    sheet.cell(rowno,columnno).fill = greenfill
    workbook.save(file)

def fillRedColor(file, sheeName, rowno, columnno, redFill=None):
    workbook=openpyxl.load_workbook(file)
    sheet= workbook[sheeName]
    greenfill = PatternFill(start_color='ff0000', end_color='ff0000',fill_type='solid')
    sheet.cell(rowno,columnno).fill = redFill
    workbook.save(file)